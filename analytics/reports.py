from datetime import datetime

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import (
    APP_NAME,
    MANAGEMENT_REPORT_FILE,
    LOW_ACTIVITY_MAX_TRANSACTIONS,
    TOP_MERCHANT_LIMIT,
)

from utils import (
    load_clean_data,
    merge_users_and_transactions,
)


def calculate_transaction_summary(transactions):
    """
    Calculate the principal transaction KPIs.
    """

    total_transactions = len(transactions)
    total_transaction_value = transactions["plan_amount"].sum()
    average_transaction_value = transactions["plan_amount"].mean()
    highest_transaction = transactions["plan_amount"].max()
    lowest_transaction = transactions["plan_amount"].min()

    return {
        "Total Transactions": total_transactions,
        "Total Transaction Value": total_transaction_value,
        "Average Transaction Value": average_transaction_value,
        "Highest Transaction": highest_transaction,
        "Lowest Transaction": lowest_transaction,
    }


def build_merchant_performance_table(users, transactions):
    """
    Build one summary row for every merchant.

    Users with no transactions are retained and assigned zero values.
    """

    transaction_summary = (
        transactions
        .groupby("user_id")
        .agg(
            total_transaction_value=("plan_amount", "sum"),
            transaction_count=("id", "count"),
            average_transaction_value=("plan_amount", "mean"),
            highest_transaction=("plan_amount", "max"),
            lowest_transaction=("plan_amount", "min"),
            last_transaction_date=("create_date", "max"),
        )
        .reset_index()
    )

    merchant_performance = users.merge(
        transaction_summary,
        left_on="id",
        right_on="user_id",
        how="left",
    )

    numeric_columns = [
        "total_transaction_value",
        "transaction_count",
        "average_transaction_value",
        "highest_transaction",
        "lowest_transaction",
    ]

    for column in numeric_columns:
        merchant_performance[column] = (
            merchant_performance[column]
            .fillna(0)
        )

    merchant_performance["transaction_count"] = (
        merchant_performance["transaction_count"]
        .astype(int)
    )

    merchant_performance = (
        merchant_performance
        .sort_values(
            by=[
                "total_transaction_value",
                "transaction_count",
            ],
            ascending=[
                False,
                False,
            ],
        )
        .reset_index(drop=True)
    )

    merchant_performance.insert(
        0,
        "rank",
        range(1, len(merchant_performance) + 1),
    )

    return merchant_performance


def build_executive_summary(users, transactions, merchant_performance):
    """
    Create a small management-facing KPI table.
    """

    transaction_summary = calculate_transaction_summary(
        transactions
    )

    merchants_with_transactions = (
        transactions["user_id"]
        .nunique()
    )

    merchants_without_transactions = (
        len(users) - merchants_with_transactions
    )

    if merchant_performance.empty:
        top_merchant = "None"
        top_merchant_value = 0
    else:
        top_merchant = merchant_performance.iloc[0]["username"]
        top_merchant_value = merchant_performance.iloc[0][
            "total_transaction_value"
        ]

    summary_rows = [
        {
            "Metric": "Application",
            "Value": APP_NAME,
        },
        {
            "Metric": "Report Generated",
            "Value": datetime.now().strftime(
                "%Y-%m-%d %H:%M"
            ),
        },
        {
            "Metric": "Total Registered Users",
            "Value": len(users),
        },
        {
            "Metric": "Users With Transactions",
            "Value": merchants_with_transactions,
        },
        {
            "Metric": "Users Without Transactions",
            "Value": merchants_without_transactions,
        },
        {
            "Metric": "Total Transactions",
            "Value": transaction_summary[
                "Total Transactions"
            ],
        },
        {
            "Metric": "Total Transaction Value",
            "Value": transaction_summary[
                "Total Transaction Value"
            ],
        },
        {
            "Metric": "Average Transaction Value",
            "Value": transaction_summary[
                "Average Transaction Value"
            ],
        },
        {
            "Metric": "Highest Transaction",
            "Value": transaction_summary[
                "Highest Transaction"
            ],
        },
        {
            "Metric": "Lowest Transaction",
            "Value": transaction_summary[
                "Lowest Transaction"
            ],
        },
        {
            "Metric": "Top Merchant",
            "Value": top_merchant,
        },
        {
            "Metric": "Top Merchant Transaction Value",
            "Value": top_merchant_value,
        },
    ]

    return pd.DataFrame(summary_rows)


def filter_user_type(merchant_performance, user_type):
    """
    Return merchants belonging to one user category.
    """

    return merchant_performance[
        merchant_performance["user_type"]
        .fillna("")
        .astype(str)
        .str.strip()
        .str.casefold()
        == user_type.strip().casefold()
    ].copy()


def build_low_activity_report(merchant_performance):
    """
    Return merchants with transaction counts at or below
    the configured MVP threshold.
    """

    low_activity = merchant_performance[
        merchant_performance["transaction_count"]
        <= LOW_ACTIVITY_MAX_TRANSACTIONS
    ].copy()

    return low_activity.sort_values(
        by=[
            "transaction_count",
            "total_transaction_value",
        ],
        ascending=[
            True,
            True,
        ],
    )


def select_report_columns(dataframe):
    """
    Keep only management-relevant merchant columns.
    """

    report_columns = [
        "rank",
        "id",
        "username",
        "full_name",
        "email",
        "phone",
        "user_type",
        "date_joined",
        "account_balance",
        "total_transaction_value",
        "transaction_count",
        "average_transaction_value",
        "highest_transaction",
        "lowest_transaction",
        "last_transaction_date",
    ]

    available_columns = [
        column
        for column in report_columns
        if column in dataframe.columns
    ]

    return dataframe[available_columns].copy()


def format_workbook(writer):
    """
    Apply simple professional formatting to all report sheets.
    """

    workbook = writer.book

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    currency_headers = {
        "account_balance",
        "total_transaction_value",
        "average_transaction_value",
        "highest_transaction",
        "lowest_transaction",
    }

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        for column_cells in worksheet.columns:
            column_letter = get_column_letter(
                column_cells[0].column
            )

            maximum_length = 0

            for cell in column_cells:
                if cell.value is not None:
                    maximum_length = max(
                        maximum_length,
                        len(str(cell.value)),
                    )

            worksheet.column_dimensions[
                column_letter
            ].width = min(maximum_length + 2, 40)

        header_positions = {
            cell.value: cell.column
            for cell in worksheet[1]
        }

        for header in currency_headers:
            column_number = header_positions.get(header)

            if column_number is None:
                continue

            for row_number in range(
                2,
                worksheet.max_row + 1,
            ):
                worksheet.cell(
                    row=row_number,
                    column=column_number,
                ).number_format = '₦#,##0.00'


def generate_management_report():
    """
    Generate the complete multi-sheet management workbook.
    """

    print("\n===== MANAGEMENT REPORT =====")

    users, transactions = load_clean_data()

    # This merge also confirms that merchant details can be
    # attached to transactions successfully.
    merge_users_and_transactions(
        users,
        transactions,
    )

    merchant_performance = (
        build_merchant_performance_table(
            users,
            transactions,
        )
    )

    executive_summary = build_executive_summary(
        users,
        transactions,
        merchant_performance,
    )

    overall_ranking = select_report_columns(
        merchant_performance
    )

    top_merchants = overall_ranking.head(
        TOP_MERCHANT_LIMIT
    )

    api_users = select_report_columns(
        filter_user_type(
            merchant_performance,
            "API",
        )
    )

    smart_earners = select_report_columns(
        filter_user_type(
            merchant_performance,
            "Smart Earner",
        )
    )

    low_activity = select_report_columns(
        build_low_activity_report(
            merchant_performance
        )
    )

    with pd.ExcelWriter(
        MANAGEMENT_REPORT_FILE,
        engine="openpyxl",
    ) as writer:

        executive_summary.to_excel(
            writer,
            sheet_name="Executive Summary",
            index=False,
        )

        top_merchants.to_excel(
            writer,
            sheet_name="Top Merchants",
            index=False,
        )

        overall_ranking.to_excel(
            writer,
            sheet_name="Overall Ranking",
            index=False,
        )

        api_users.to_excel(
            writer,
            sheet_name="API Users",
            index=False,
        )

        smart_earners.to_excel(
            writer,
            sheet_name="Smart Earners",
            index=False,
        )

        low_activity.to_excel(
            writer,
            sheet_name="Low Activity",
            index=False,
        )

        transactions.to_excel(
            writer,
            sheet_name="Clean Transactions",
            index=False,
        )

        format_workbook(writer)

    print(
        "\n✅ Management report generated successfully."
    )

    print(
        f"Output: {MANAGEMENT_REPORT_FILE}"
    )

    print("\nSheets created:")
    print("- Executive Summary")
    print("- Top Merchants")
    print("- Overall Ranking")
    print("- API Users")
    print("- Smart Earners")
    print("- Low Activity")
    print("- Clean Transactions")

    return MANAGEMENT_REPORT_FILE


if __name__ == "__main__":
    generate_management_report()